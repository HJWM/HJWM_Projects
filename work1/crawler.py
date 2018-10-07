# -*-coding:utf-8 -*-

# 豆瓣网爬书参考代码 日期:2018-09-24 参考代码
import sys
import time
from urllib.parse import quote
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
import lxml

#User Agents
hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; enUS; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
{'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/ 535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


#*********爬书虫函数**********
def book_spider(book_tag):
  page_num = 0 # 页数
  book_list = []  # 爬到的书籍列表
  try_times = 0 # 尝试次数

  while (page_num < 1):  #只爬一页
     # url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # 测试用
     #每类爬15本书
     url = 'http://www.douban.com/tag/' + quote(book_tag) + '/book?start=' + str(page_num*15)
     time.sleep(np.random.rand() * 1)
    #开始爬取
     try:
        req = requests.request('get',url, headers=hds[page_num % len(hds)])
        source_code =req.text
        plain_text = str(source_code)
     except Exception as e:
        print(e)
        continue

     ##测试 防止IP被禁用
     #source_code = requests.get(url)
     #plain_text = source_code.text

     #使用BS4 解析网页内容
     soup = BeautifulSoup(plain_text,"lxml")
     list_soup = soup.find('div', {'class': 'mod book-list'})

     try_times += 1
     if list_soup == None and try_times < 200:
        continue
     elif list_soup == None or len(list_soup) <= 1:
        break  # 200 次请求后 任然没有数据就直接退出

    #分析书籍列表
     for book_info in list_soup.ﬁndAll('dd'):
        #标题
        title = book_info.ﬁnd('a', {'class': 'title'}).string.strip()
        #描述
        desc = book_info.ﬁnd('div', {'class': 'desc'}).string.strip()
        desc_list = desc.split('/')
        #url 地址
        book_url = book_info.ﬁnd('a', {'class': 'title'}).get('href')
        try:
           author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
        except:
           author_info = '作者/译者： 暂无'
        try:
           pub_info = desc_list[-3]
        except:
           pub_info = '未知'
        try:
           date_info = desc_list[-2]
        except:
           date_info = '0000-00-00'
        try:
           price_info =desc_list[-1].strip('元')  #去除"元"这个字
        except:
           price_info = '0.0'
        try:
           rating = book_info.ﬁnd('span',{'class': 'rating_nums'}).string.strip()
        except:
           rating = '0.0'
        try:
           people_num = get_people_num(book_url)
           people_num = people_num.strip('人评价')  #去除"人评价"这几个字
        except:
           people_num = '0'
           # 跟踪输出
        print([title, rating, people_num, author_info, pub_info, date_info,price_info])
        book_list.append([title, rating, people_num, author_info, pub_info, date_info,price_info])
        if (len(book_list) > 20):
           break
        try_times = 0   # 获取无效信息设置成 0

     page_num += 1
     # print(page_num)
     print('从第{}页获取网页信息'.format(str(page_num)))
  return book_list

#*************获取评价人数************
def get_people_num(url):
   # url='http://book.douban.com/subject/6082808/?from=tag_all' # 地址栏测试用
   try:
       req = requests.request('get',url, headers=hds[np.random.randint(0, len(hds))])
       source_code = req.text
       plain_text = str(source_code)
   except Exception as e:
       print(e)
   # 获取评价人数
   soup = BeautifulSoup(plain_text, 'lxml')
   people_num = soup.ﬁnd('div', {'class': 'rating_sum'}).ﬁndAll('span')[1].string.strip()
   return people_num


#**************分类爬书籍***********
def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
       book_list = book_spider(book_tag)
       book_list = sorted(book_list, key=lambda x: x[1], reverse=True) #利用第二个参数进行升序排序
       book_lists.append(book_list)
    return book_lists


#****************数据保存到excel中*********
def print_book_lists_excel(book_lists, book_tag_lists):
   wb = Workbook()
   ws = []
   for i in range(len(book_tag_lists)):
       ws.append(wb.create_sheet(title=book_tag_lists[i]))
   for i in range(len(book_tag_lists)):
       ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社', '出版日期', '价格'])
       count = 1
       for bl in book_lists[i]:
           ws[i].append([count, bl[0],float(bl[1]), int(bl[2]), bl[3], bl[4], bl[5], bl[6]])
           count += 1
   # save_path = 'book_list'  # 保存
   # for i in range(len(book_tag_lists)):
   #     save_path += ('-' + book_tag_lists[i])
   #     save_path += '.xlsx'
   #     wb.save(save_path)
   save_path='book_list2.xlsx'
   wb.save(save_path)


""" 主函数 """
if __name__ == '__main__':
  # 定义书要爬籍类型别
  book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
  book_lists = do_spider(book_tag_lists)
  print_book_lists_excel(book_lists, book_tag_lists)
  #  book_tag_lists = ['传记','哲学','编程','创业','理理财','社会学','佛教']
  #  book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
  #  book_tag_lists = ['计算机','机器器学习','linux','android','数据库','互联 ⽹网']
  #  book_tag_lists = ['数学']
  #  book_tag_lists = ['摄影','设计','音乐','旅⾏','教育','成长','情感','育儿','健康','养生']
  #  book_tag_lists = ['商业','理财','管理']
  #  book_tag_lists = ['名著']
  #  book_tag_lists = ['科普','经典','生活','心灵','文学']
  #  book_tag_lists = ['科幻','思维','⾦融']
  # book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']